"""
封装一个读取用例的excel类
# 实现读取用例数据
# 实现写入数据的功能
"""
from collections import namedtuple

import openpyxl
from openpyxl.styles import Font


class Case:
    def __repr__(self):
        string = '\n In __repr__：'
        string += '\n' + repr(self.__dict__)
        return string


class ReadYaml:
    """
    读取yaml数据
    """

    def __init__(self, file_name):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        """
        pass


class ReadExcel(object):
    """
    读取excel数据
    """

    def __init__(self, file_name, sheet_name):
        """
        这个是用例初始化读取对象的
        :param file_name:  文件名字  -->  str
        :param sheet_name: 表单名字  -->  str
        """

        # 先将多余的行列预处理删除掉
        self.wb = openpyxl.load_workbook(file_name)
        self.sheet = self.wb[sheet_name]
        self.max_row = self.sheet.max_row
        self.max_column = self.sheet.max_column
        for row in list(self.sheet.rows)[::-1]:
            if row[0].value is None:
                self.sheet.delete_rows(row[0].row, 1)
        for column in list(self.sheet.columns)[::-1]:
            if column[0].value is None:
                self.sheet.delete_cols(column[0].column, 1)
        self.file_name = file_name
        self.sheet_name = sheet_name

        self.save()
        self.close()

    def __del__(self):
        self.close()

    @staticmethod
    def create_new_workbook(filename):
        if ".xlsx" not in filename:
            filename = filename + ".xlsx"
        wb = openpyxl.Workbook()
        wb.create_sheet(title="Sheet1", index=0)
        wb.save(filename)

    @staticmethod
    def get_all_sheets(filename):
        wb = openpyxl.load_workbook(filename)
        return wb.sheetnames

    def clear_sheet(self):
        self.wb.remove(self.wb[self.sheet_name])
        self.wb.create_sheet(self.sheet_name, 0)
        self.save()

    def clear_sheet_except_title(self):
        for row in list(self.sheet.rows)[:0:-1]:
            self.sheet.delete_rows(row[0].row, 1)
        self.save()

    def read_data_line(self):
        """
        按行读取数据
        :return:  返回一个列表，列表中每个元素为一条用例
        """
        self.open()
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例
        cases = []
        # 遍历出所有测试用例
        for case in rows_data[1:]:
            # data用例临时存放用例数据
            data = []
            for cell in case:
                # 判断该单元格数据是否为字符串类型，
                data.append(cell.value)
            # 将表头和，该条数据内容，打包成一个字典，放入cases中
            case_data = dict(list(zip(titles, data)))
            cases.append(case_data)
        return cases

    def read_data_obj(self):
        """
        按行读取数据，表单所有数据
        每个用例存储在一个对象中
        :return: 返回一个列表，列表中每个元素为一个用例对象
        """
        self.open()
        # 按行获取数据转换成列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        titles.append("max_column")
        # 定义一个空列表用来存储所有的用例
        if 'result' not in titles:
            self.w_data(1, self.r_max()[1] + 1, 'result')
        cases = []
        for case in rows_data[1:]:
            # 创建一个Cases类的对象，用来保存用例数据，
            case_obj = Case()
            # data用例临时存放用例数据
            data = []
            # 判断该单元格是否为字符串类型，
            for cell in case:
                data.append(cell.value)
            data.append(len(case))
            # 将该条数据放入cases中
            case_data = list(zip(titles, data))
            for i in case_data:
                if i[0] == 'result' or i[0] is None:
                    continue
                setattr(case_obj, i[0], i[1])
            setattr(case_obj, 'row', case[0].row)
            cases.append(case_obj)

        return cases

    def r_data_from_line(self, list1):
        """
        按list中行读取，表单所有该行数据
        每个用例存储在一个字典中
        :param list1:  -->  要求读取的行编号列表，eg:[1,3,5]则会读取第1,3,5条用例
        :return: 包含所有用例字典的列表list
        """
        # list1 参数为一个列表，传入的是指定读取数据的行,比如[1,2,3]
        # 每一行[1,3,5]列的数据，读取出来就作为一条测试用例，放在字典中
        # 所有的用例放在列表中并且进行返回
        self.open()
        rows_data = list(self.sheet.rows)
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        cases = []
        for k in list1:
            temp = []
            for cell in rows_data[k]:
                temp.append(cell.value)
            case_data = dict(zip(titles, temp))
            cases.append(case_data)
        return cases

    def r_data_obj_from_line(self, list1):
        """
        按list中行读取，表单所有该行数据
        每个用例存储在一个对象中
        :param list1: list1  -->   要求读取的行编号列表，eg:[1,3,5]则会读取第1,3,5条用例
        :return: 包含所有用例对象的列表list
        """
        # list1 参数为一个列表，传入的是指定读取数据的列,比如[1,2,3]
        # 每一行[1,3,5]列的数据，读取出来就作为一条测试用例，放在对象中属性中
        # 所有的用例对象放在列表中并且进行返回
        self.open()
        rows_data = list(self.sheet.rows)
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        case_all = []
        for k in list1:
            Cases = namedtuple('Cases', ['%s' % x for x in titles])
            case_one = []
            for cell in rows_data[k]:
                case_one.append(cell.value)
            case_obj = Cases(*case_one)
            case_all.append(case_obj)
        return case_all

    def r_data_from_column(self, list1):
        """
        按list中行读取，表单所有该行数据
        每个用例存储在一个对象中
        :param list1: list1  -->   要求读取的列编号列表，eg:[1,3,5]则会读取所有用例的第1,3,5列
        :return: 包含所有用例对象的列表list
        """
        titles = []
        case_all = []
        for row in range(1, self.sheet.max_row + 1):
            case_one = []
            if row == 1:
                for column in list1:
                    titles.append(self.sheet.cell(row, column).value)
            else:
                for column in list1:
                    case_one.append(self.sheet.cell(row, column).value)
                case_all.append(dict(zip(titles, case_one)))
        return case_all

    def r_data_obj_from_column(self, list1):
        """
        按list中行读取，表单所有该行数据
        每个用例存储在一个对象中
        :param list1: list1  -->   要求读取的列编号列表，eg:[1,3,5]则会读取所有用例的第1,3,5列
        :return: 包含所有用例对象的列表list
        """
        self.open()
        if list1 is None:
            return self.read_data_obj()
        titles = []
        case_all = []
        for row in range(1, self.sheet.max_row + 1):
            case_one = []
            if row == 1:
                for column in list1:
                    titles.append(self.sheet.cell(row, column).value)
            else:
                for column in list1:
                    case_one.append(self.sheet.cell(row, column).value)
                cases = namedtuple('Cases', titles)
                case_obj = cases(*case_one)
                case_all.append(case_obj)
        return case_all

    def open(self):
        self.wb = openpyxl.load_workbook(self.file_name)
        self.sheet = self.wb[self.sheet_name]
        self.save()

    def close(self):
        self.wb.close()

    def w_data(self, row, column, data):
        self.open()
        self.sheet.cell(row, column, data)
        self.save()
        self.close()

    def set_column_width(self, column, width):
        self.sheet.column_dimensions[column].width = width
        self.save()

    def set_font(self, row, column, Font=Font(u'宋体', size=11, bold=True, color='000000')):
        cell = chr(ord("A") + column - 1) + str(row)
        self.sheet[cell].font = Font
        self.save()

    def w_data_origin(self, row, column, data):
        self.sheet.cell(row, column, data)

    def save(self):
        self.wb.save(self.file_name)

    def r_max(self):
        """
        0 为 最大行数， 1 为最大列数
        :return:
        """
        return self.max_row, self.max_column


if __name__ == '__main__':
    r = ReadExcel('../TestCases/data/LoginTestData.xlsx', 'Login')
    # print('---------------------------------------------------')
    # data = r.r_data_from_colunm([1, 2, 3])
    # print(data)
    #
    # print('---------------------------------------------------')
    # data = r.r_data_obj_from_column([1, 2, 3])
    # print(data)
    res = r.read_data_obj()
    print(res)
    # r.clear_sheet()
    # r.set_column_width("D", 14)
    # FONT = Font(u'宋体', size=12, bold=True, color='000000')
    # r.sheet.column_dimensions["A"].font = FONT
    # r.set_font("A3", FONT)
    # r.save()
    # ReadExcel.create_new_workbook("july.xlsx")
